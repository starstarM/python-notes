from openpyxl import load_workbook,Workbook

open_excel = load_workbook("python.xlsx")

sheet_name = open_excel.sheetnames
sheet = open_excel.get_sheet_by_name(sheet_name[0])

anchors =[]
for i in range(1, sheet.max_row+1):
    name  = sheet["A{0}".format(i)].value
    num = sheet["B{0}".format(i)].value
    
    anchor = {"location":name, "number":num}
    anchors.append(anchor)
def sort_seed(anchor):
    return anchor["number"]
sort_an = sorted(anchors, key=sort_seed, reverse=False)

wb = Workbook()
sheet1 = wb.active
for i in range(1, sheet.max_row+1):
    sheet1['A{0}'.format(i)] = sort_an[i-1]["location"]
    sheet1['B{0}'.format(i)] = sort_an[i-1]["number"]

wb.save('python.xlsx')


